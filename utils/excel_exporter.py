# utils/excel_exporter.py
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, Border, Side, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing import image
from datetime import datetime
import io
from pathlib import Path

class LogisticsExcelExporter:
    """
    Excel exporter that creates formatted reports matching the exact layout
    and styling of the logistics cost calculation template.
    """
    
    def __init__(self):
        # Define color schemes matching the template
        self.colors = {
            'header_blue': 'D6E3F0',      # Light blue for headers
            'section_gray': 'F2F2F2',      # Light gray for section headers
            'summary_yellow': 'FFF2CC',    # Light yellow for summary
            'border_gray': '8E8E8E',       # Gray for borders
            'text_black': '000000',        # Black text
            'text_blue': '0066CC'          # Blue text for headers
        }
        
        # Define fonts
        self.fonts = {
            'title': Font(name='Arial', size=16, bold=True, color=self.colors['text_black']),
            'header': Font(name='Arial', size=12, bold=True, color=self.colors['text_blue']),
            'subheader': Font(name='Arial', size=10, bold=True, color=self.colors['text_black']),
            'normal': Font(name='Arial', size=10, color=self.colors['text_black']),
            'normal_bold': Font(name='Arial', size=10, bold=True, color=self.colors['text_black']),  # Added for bold normal text
            'small': Font(name='Arial', size=9, color=self.colors['text_black'])
        }
        
        # Define fills
        self.fills = {
            'header_blue': PatternFill(start_color=self.colors['header_blue'], 
                                     end_color=self.colors['header_blue'], 
                                     fill_type='solid'),
            'section_gray': PatternFill(start_color=self.colors['section_gray'], 
                                      end_color=self.colors['section_gray'], 
                                      fill_type='solid'),
            'summary_yellow': PatternFill(start_color=self.colors['summary_yellow'], 
                                        end_color=self.colors['summary_yellow'], 
                                        fill_type='solid')
        }
        
        # Define borders
        thin_border = Side(border_style="thin", color=self.colors['border_gray'])
        self.border = Border(top=thin_border, left=thin_border, 
                           right=thin_border, bottom=thin_border)
    
    def create_logistics_report(self, calculation_result, plant_id="1051", version="1.5.5", created_by="System"):
        """
        Create a formatted Excel report matching the logistics cost calculation template.
        
        Args:
            calculation_result: Dictionary containing calculation results
            plant_id: Plant identifier
            version: Report version
            created_by: Creator name
        
        Returns:
            BytesIO: Excel file in memory
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Logistics Cost Calculation"
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        
        row = 1
        
        # Title and Header Information
        row = self._add_header_section(ws, row, calculation_result, plant_id, version, created_by)
        
        # Summary Section
        row = self._add_summary_section(ws, row, calculation_result)
        
        # General Information
        row = self._add_general_information(ws, row, calculation_result)
        
        # Material Information
        row = self._add_material_information(ws, row, calculation_result)
        
        # Supplier Information
        row = self._add_supplier_information(ws, row, calculation_result)
        
        # Operations Information
        row = self._add_operations_information(ws, row, calculation_result)
        
        # Packaging Information
        row = self._add_packaging_information(ws, row, calculation_result)
        
        # Calculation Details
        row = self._add_calculation_section(ws, row, calculation_result)
        
        # Transport Information
        row = self._add_transport_section(ws, row, calculation_result)
        
        # Warehouse and Additional Costs
        row = self._add_warehouse_section(ws, row, calculation_result)
        
        # Final Summary
        row = self._add_final_summary(ws, row, calculation_result)
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
    
    def _add_header_section(self, ws, start_row, data, plant_id, version, created_by):
        """Add the header section with title and metadata"""
        row = start_row
        
        # Main title
        ws.merge_cells(f'A{row}:C{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = "Logistic cost calculation"
        title_cell.font = self.fonts['title']
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = self.fills['header_blue']
        title_cell.border = self.border
        row += 1
        
        # Metadata
        metadata = [
            ("Plant", plant_id),
            ("Date of creation", datetime.now().strftime("%d.%m.%Y")),
            ("Created by", created_by),
            ("Version", version)
        ]
        
        for label, value in metadata:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = self.fonts['normal']
            ws[f'B{row}'].font = self.fonts['normal']
            row += 1
        
        return row + 1
    
    def _add_summary_section(self, ws, start_row, data):
        """Add the summary section with key cost components"""
        row = start_row
        
        # Summary header
        ws.merge_cells(f'A{row}:C{row}')
        header_cell = ws[f'A{row}']
        header_cell.value = "Summary"
        header_cell.font = self.fonts['header']
        header_cell.fill = self.fills['summary_yellow']
        header_cell.border = self.border
        row += 1
        
        # Calculate summary values
        customs_cost = data.get('customs_cost_per_piece', 0)
        transport_cost = data.get('transport_cost_per_piece', 0) + data.get('co2_cost_per_piece', 0)
        packaging_cost = data.get('packaging_cost_per_piece', 0)
        ald_cost = data.get('warehouse_cost_per_piece', 0) + data.get('repacking_cost_per_piece', 0)
        total_cost_per_piece = data.get('total_cost_per_piece', 0)
        total_annual_cost = data.get('total_annual_cost', 0)
        
        summary_items = [
            ("customs (for part and transport in third country)", f"{customs_cost:.3f} €"),
            ("Transport (transport + co2)", f"{transport_cost:.3f} €"),
            ("Packaging", f"{packaging_cost:.3f} €"),
            ("ALD (warehousing + re-packaging)", f"{ald_cost:.3f} €"),
            ("Logistic costs per pcs", f"{total_cost_per_piece:.3f} €"),
            ("Logistic costs per year", f"{total_annual_cost:.2f} €")
        ]
        
        for label, value in summary_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = self.fonts['normal']
            ws[f'B{row}'].font = self.fonts['normal']
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        return row + 1
    
    def _add_general_information(self, ws, start_row, data):
        """Add general information section"""
        row = start_row
        
        # Section header
        ws.merge_cells(f'A{row}:C{row}')
        header_cell = ws[f'A{row}']
        header_cell.value = "General information"
        header_cell.font = self.fonts['header']
        header_cell.fill = self.fills['section_gray']
        header_cell.border = self.border
        row += 1
        
        return row 
    
    def _add_material_information(self, ws, start_row, data):
        """Add material information section"""
        row = start_row
        
        # Section header
        ws.merge_cells(f'A{row}:C{row}')
        header_cell = ws[f'A{row}']
        header_cell.value = "Material Information"
        header_cell.font = self.fonts['subheader']
        header_cell.fill = self.fills['section_gray']
        header_cell.border = self.border
        row += 1
        
        material_info = [
            ("Project Name", data.get('Project Name', '')),
            ("Material No.", data.get('material_id', '')),
            ("Material Description", data.get('material_desc', '')),
            ("Annual volume (average)", f"{data.get('Annual Volume', 0):,}"),
            ("SOP", data.get('SOP', ''))
        ]
        
        for label, value in material_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = self.fonts['normal']
            ws[f'B{row}'].font = self.fonts['normal']
            row += 1
        
        return row + 1
    
    def _add_supplier_information(self, ws, start_row, data):
        """Add supplier information section"""
        row = start_row
        
        # Section header
        ws.merge_cells(f'A{row}:C{row}')
        header_cell = ws[f'A{row}']
        header_cell.value = "Supplier Information"
        header_cell.font = self.fonts['subheader']
        header_cell.fill = self.fills['section_gray']
        header_cell.border = self.border
        row += 1
        
        supplier_info = [
            ("Vendor ID", data.get('supplier_id', '')),
            ("Vendor Name", data.get('supplier_name', '')),
            ("Vendor Country", data.get('Vendor Country', '')),
            ("City of manufacture", data.get('City of Manufacture', '')),
            ("Vendor ZIP", data.get('Vendor ZIP', '')),
            ("Deliveries per month", data.get('Deliveries per Month', ''))
        ]
        
        for label, value in supplier_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = self.fonts['normal']
            ws[f'B{row}'].font = self.fonts['normal']
            row += 1
        
        return row + 1
    
    def _add_operations_information(self, ws, start_row, data):
        """Add operations information section"""
        row = start_row
        
        # Section header
        ws.merge_cells(f'A{row}:C{row}')
        header_cell = ws[f'A{row}']
        header_cell.value = "Operations Information"
        header_cell.font = self.fonts['subheader']
        header_cell.fill = self.fills['section_gray']
        header_cell.border = self.border
        row += 1
        
        operations_info = [
            ("Incoterm Code", data.get('Incoterm code', '')),
            ("Incoterm Named Place", data.get('Incoterm Named Place', '')),
            ("MOQ*", data.get('MOQ', '')),
            ("Call-off transfer type", data.get('Call-off transfer type', '')),
            ("Lead time (d)", data.get('Lead time (d)', '')),
            ("Sub-supplier used?", data.get('Sub-Supplier Used', ''))
        ]
        
        for label, value in operations_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = self.fonts['normal']
            ws[f'B{row}'].font = self.fonts['normal']
            row += 1
        
        return row + 1
    
    def _add_packaging_information(self, ws, start_row, data):
        """Add packaging information section"""
        row = start_row
        
        # Section header
        ws.merge_cells(f'A{row}:C{row}')
        header_cell = ws[f'A{row}']
        header_cell.value = "Packaging Information"
        header_cell.font = self.fonts['subheader']
        header_cell.fill = self.fills['section_gray']
        header_cell.border = self.border
        row += 1
        
        packaging_info = [
            ("Packaging type", data.get('packaging_type', '')),
            ("Filling degree per box", f"{data.get('Filling degree per box', 0):,}"),
            ("Filling degree per pallet", f"{data.get('Filling degree per pallet', 0):,}"),
            ("Special packaging required", data.get('Special packaging required', ''))
        ]
        
        for label, value in packaging_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = self.fonts['normal']
            ws[f'B{row}'].font = self.fonts['normal']
            row += 1
        
        return row + 1
    
    def _add_calculation_section(self, ws, start_row, data):
        """Add detailed calculation section"""
        row = start_row
        
        # Section header
        ws.merge_cells(f'A{row}:C{row}')
        header_cell = ws[f'A{row}']
        header_cell.value = "Calculation"
        header_cell.font = self.fonts['header']
        header_cell.fill = self.fills['summary_yellow']
        header_cell.border = self.border
        row += 1
        
        # Price and packaging costs - BOLD
        price_pcs = data.get('Price (Pcs)', 0)
        packaging_cost_pcs = data.get('packaging_cost_per_piece', 0)
        
        calc_items = [
            ("Price (pcs)", f"{price_pcs:.3f} €"),
            ("Packaging cost (pcs)", f"{packaging_cost_pcs:.3f} €")
        ]
        
        for label, value in calc_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = self.fonts['normal_bold']  # Changed to bold
            ws[f'B{row}'].font = self.fonts['normal']
            row += 1
        
        # Packaging breakdown
        plant_cost = data.get('packaging_cost_plant', 0)
        coc_cost = data.get('packaging_cost_coc', 0)
        
        ws[f'A{row}'] = "Plant"
        ws[f'B{row}'] = f"{plant_cost:.2f} €"
        row += 1
        
        ws[f'A{row}'] = "CoC"
        ws[f'B{row}'] = f"{coc_cost:.2f} €"
        row += 1
        
        # Packaging Loop - BOLD
        packaging_loop = data.get('Packaging Loop', 0)
        ws[f'A{row}'] = "Packaging Loop"
        ws[f'B{row}'] = f"{packaging_loop} days"
        ws[f'A{row}'].font = self.fonts['normal_bold']  # Changed to bold
        ws[f'B{row}'].font = self.fonts['normal']
        row += 1
        
        # Loop details with indentation
        loop_items = [
            ("goods receipt", data.get('goods_receipt', 0)),
            ("stock", data.get('stock_raw_materials', 0)),
            ("production", data.get('production', 0)),
            ("empties", data.get('empties_return', 0)),
            ("cleaning", data.get('cleaning', 0)),
            ("dispatch", data.get('dispatch', 0)),
            ("empties transit KB to supplier", data.get('empties_transit_kb_to_supplier', 0)),
            ("empties receipt", data.get('empties_receipt_at_supplier', 0)),
            ("empties stock", data.get('empties_in_stock_supplier', 0)),
            ("production", data.get('production_contrary_loop', 0)),
            ("stock finished parts", data.get('stock_finished_parts', 0)),
            ("dispatch", data.get('dispatch_finished_parts', 0)),
            ("transit supplier to KB", data.get('transit_supplier_to_plant', 0))
        ]
        
        for label, value in loop_items:
            ws[f'A{row}'] = f" {label}"  # Indentation with space
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = self.fonts['small']
            ws[f'B{row}'].font = self.fonts['small']
            row += 1
        
        # Other cost components - BOLD
        other_costs = [
            ("Repacking cost (pcs)", f"{data.get('repacking_cost_per_piece', 0):.3f} €"),
            ("Customs cost (pcs)", f"{data.get('customs_cost_per_piece', 0):.3f} €"),
            ("Duty rate (% of pcs price)", f"{data.get('Duty Rate (% Of pcs price)', 0):.1f}%"),
            ("Transport cost (pcs)", f"{data.get('transport_cost_per_piece', 0):.3f} €")
        ]
        
        for label, value in other_costs:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            # Make these labels bold
            if label in ["Repacking cost (pcs)", "Transport cost (pcs)", "customs cost (pcs)"]:
                ws[f'A{row}'].font = self.fonts['normal_bold']
            else:
                ws[f'A{row}'].font = self.fonts['normal']
            ws[f'B{row}'].font = self.fonts['normal']
            row += 1
        
        return row  
    
    def _add_transport_section(self, ws, start_row, data):
        """Add transport section"""
        row = start_row
        
        transport_info = [
            ("Transport type", data.get('Transport type', '')),
            ("Pallets per delivery", data.get('pallets_per_delivery', 0)), 
            ("Transportation cost per LU", f"{data.get('Transport cost per LU', 0):.2f} €"),
            ("annual CO2 cost (pcs)", f"{data.get('co2_cost_per_piece', 0):.3f} €")
        ]
        
        for label, value in transport_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            # Make annual CO2 cost bold
            if label == "annual CO2 cost (pcs)":
                ws[f'A{row}'].font = self.fonts['normal_bold']
            else:
                ws[f'A{row}'].font = self.fonts['normal']
            ws[f'B{row}'].font = self.fonts['normal']
            row += 1
        
        return row 
    
    def _add_warehouse_section(self, ws, start_row, data):
        """Add warehouse and additional costs section"""
        row = start_row
        
        warehouse_info = [
            ("Warehouse cost (pcs)", f"{data.get('warehouse_cost_per_piece', 0):.3f} €"),
            ("Safety stock (No. pal)", f"{data.get('safety_stock_no_pal', 0):.0f}"),
            ("Excessive inventory cost (pcs)", "0,000 €"),  # Default as shown
            ("Total Inventory cost (parts in WH)", "0,000 €"),  # Default as shown
            ("Inventory Interest", "0,000 €"),  # Default as shown
            ("other / additional cost (pcs)", f"{data.get('additional_cost_per_piece', 0):.3f} €")
        ]
        
        # List of labels that should be bold
        bold_labels = [
            "Warehouse cost (pcs)",
            "Excessive inventory cost (pcs)",
            "Total Inventory cost (parts in WH)",
            "Inventory Interest",
            "other / additional cost (pcs)"
        ]
        
        for label, value in warehouse_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            # Make specified labels bold
            if label in bold_labels:
                ws[f'A{row}'].font = self.fonts['normal_bold']
            else:
                ws[f'A{row}'].font = self.fonts['normal']
            ws[f'B{row}'].font = self.fonts['normal']
            row += 1
        
        return row + 1
    
    def _add_final_summary(self, ws, start_row, data):
        """Add final summary section"""
        row = start_row
        
        total_cost_per_piece = data.get('total_cost_per_piece', 0)
        total_annual_cost = data.get('total_annual_cost', 0)
        
        final_summary = [
            ("Logistics cost per pcs", f"{total_cost_per_piece:.3f} €"),
            ("Logistics cost per year", f"{total_annual_cost:.2f} €")
        ]
        
        for label, value in final_summary:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = self.fonts['subheader']
            ws[f'B{row}'].font = self.fonts['subheader']
            ws[f'A{row}'].fill = self.fills['summary_yellow']
            ws[f'B{row}'].fill = self.fills['summary_yellow']
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        return row + 1